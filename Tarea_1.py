#Lo primero que hago es importar las librerías necesarias, localizar el excel para guardar los datos y crear el bucle while True para que nos repita el menú hasta que decidamos salir:
import os
import getpass # Sirve para ocultar el número en el modo de 2 jugadores
import random # Nos ayuda a crear un número aleatorio
import openpyxl # Necesario para poder guardar las estadísticas en el excel
Ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Estadísticas.xlsx")
if os.path.exists(Ruta):
    # Si el archivo ya existe, lo abrimos
    Excel = openpyxl.load_workbook(Ruta)
    Hoja = Excel[Excel.sheetnames[0]]  # Usamos la primera hoja
else:
    # Si el archivo no existe, lo creamos
    Excel = openpyxl.Workbook()
    Hoja = Excel.active
    Hoja.title = "Hoja1"
    
    # Guardamos inmediatamente para que el archivo se cree físicamente
    Excel.save(Ruta)

# Antes de eso comprobamos las etiquetas de cada valor en el encabezado y si no están creadas las creamos en la primera fila:

encabezados = [Hoja.cell(row=1, column=i).value for i in range(1, 5)]
if encabezados != ["Nombre", "Resultado", "Intentos", "Dificultad"]:
    Hoja.cell(row=1, column=1, value="Nombre")
    Hoja.cell(row=1, column=2, value="Resultado")
    Hoja.cell(row=1, column=3, value="Intentos")
    Hoja.cell(row=1, column=4, value="Dificultad")
    Excel.save(Ruta)  # Guarda el encabezado si lo has creado

    
# También necesitamos crear una función para identificar la información de cada jugador en el excel:    
    
def Registrar_jugador(Nombre, Resultado, Intentos, Dificultad):
    fila=Hoja.max_row + 1
    Hoja.cell(row=fila, column=1, value= Nombre)
    Hoja.cell(row=fila, column=2, value= Resultado)
    Hoja.cell(row=fila, column=3, value= Intentos)
    Hoja.cell(row=fila, column=4, value= Dificultad)
    Excel.save (Ruta)
    
while True:
    print ("Bienvenido, elija un modo de juego")
    Modos_de_juego= "1. Partida modo solitario\n2. Partida 2 Jugadores\n3. Estadística\n4. Salir"
    Dificultad_juego = "1. Fácil (20 intentos)\n2. Medio (12 intentos)\n3. Difícil (5 intentos)"
    print (Modos_de_juego)
    def Valida1(minimo, maximo):
        while True:
            try:
                Variable1=int(input("Elija una de las 4 opciones: "))
                if Variable1 < minimo or Variable1 > maximo:
                    print("Esta opción no es válida, por favor indique uno de las 4 opciones: ")
                else:   
                    return Variable1
            except ValueError:
                print("Por favor, prueba a introducir un número del 1 al 4 para elegir una opción del menú")
    Variable1= Valida1(1,4)
    
    # Una vez metido el menú, pongo las 4 opciones detalladas con todos sus requisitos:
    
    if Variable1 == 1:
        Nombre1 = input("Has elegido el modo solitario, por favor escriba su nombre: ")
        print ("Hola " + Nombre1 + ",¿En qué dificultad quieres jugar?")
        print (Dificultad_juego)
        def Valida2(minimo, maximo):
            while True:
                try:
                    Variable2=int(input("Elija una de las 3 dificultades: "))
                    if Variable2 < minimo or Variable2 > maximo:
                        print("Esta opción no es válida, por favor indique uno de las 3 opciones: ")
                    else:
                        return Variable2
                except ValueError:
                    print("Por favor, prueba a introducir un número del 1 al 3 para elegir la dificultad")
        Variable2= Valida2(1,3)
        if Variable2 == 1:
            print ("Has elegido la dificultad Fácil") 
            max_intentos=20
            nombre_dificultad = "Fácil"  #Nos sirve a la hora de crear el excel poder identificar la dificultad como nombre      
        if Variable2 == 2:
            print ("Has elegido la dificultad Medio")
            max_intentos=12
            nombre_dificultad = "Medio"
        if Variable2 == 3:
            print ("Has elegido la dificultad Difícil")
            max_intentos=5
            nombre_dificultad = "Difícil"
        Numero_adivinar= random.randint(1,1000)
        def Juego1():
            intentos=0
            while intentos < max_intentos:  # Usamos el try y except para asegurarnos de que se introduce un número y no hay error
                try:
                    Numero_insertado=input("Prueba suerte con un número del 1 al 1000: ").strip() #No es un texto
                    if not Numero_insertado:
                        print("No has introducido un número. Inténtalo de nuevo con un número del 1 al 1000.")
                        continue
                    Num_Jugador= int(Numero_insertado) #Ahora si tiene que ser número
                    if Num_Jugador > Numero_adivinar:
                        intentos +=1
                        print ("Este número es mayor, pruebe con uno más pequeño")
                    elif Num_Jugador < Numero_adivinar:
                        intentos +=1
                        print ("Este número es más pequeño, pruebe con uno mayor")
                    else:
                        intentos +=1
                        print(print(f"Enhorabuena {Nombre1}, ¡Has ganado! El número era el {Numero_adivinar} y lo has conseguido en el intento número {intentos} de {max_intentos} posibles"))
                        Registrar_jugador(Nombre1, "Ganador", intentos, nombre_dificultad)
                        break
                except ValueError: #Cualquier caracter extraño
                    print("No has introducido un número. Inténtalo de nuevo con un número del 1 al 1000")
            else:
                print(f"Lo siento {Nombre1}, no has conseguido adivinar el número {Numero_adivinar}, en {max_intentos} intentos.")
                Registrar_jugador(Nombre1, "Perdedor", max_intentos, nombre_dificultad)
        Juego1()
    
# Se configura el modo para 2 jugadores, de forma parecida al modo en solitario:
        
    if Variable1 == 2:
        print ("Has elegido el modo para 2 jugadores")
        Nombre1 = input("Escribe el nombre del jugador 1:  ")
        Nombre2 = input("Escribe el nombre del jugador 2:  ")
        print ("Hola " + Nombre1 + " y " + Nombre2 + ",¿En qué dificultad queréis jugar?")
        print (Dificultad_juego)
        def Valida2(minimo, maximo):
            while True:
                try:
                    Variable2=int(input("Elija una de las 3 dificultades: "))
                    if Variable2 < minimo or Variable2 > maximo:
                        print("Esta opción no es válida, por favor indique uno de las 3 opciones: ")
                    else:
                        return Variable2
                except ValueError:
                    print("Por favor, prueba a introducir un número del 1 al 3 para elegir la dificultad")
        Variable2= Valida2(1,3)
        if Variable2 == 1:
            print ("Has elegido la dificultad Fácil") 
            max_intentos=20      
            nombre_dificultad = "Fácil"
        if Variable2 == 2:
            print ("Has elegido la dificultad Medio")
            max_intentos=12
            nombre_dificultad = "Medio"
        if Variable2 == 3:
            print ("Has elegido la dificultad Difícil")
            max_intentos=5
            nombre_dificultad = "Difícil"
        while True:     # Hacemos este bucle para evitar que se escriba un texto o un caracter no válido cuando usamos la librería getpass:
            try:
                Numero_adivinar= getpass.getpass(prompt="Teclee un número entre el 1-1000 para que el jugador 2 lo adivine: ") #No es texto
                if not Numero_adivinar:
                    print ("No has introducido un número. Inténtalo de nuevo con un número del 1 al 1000.")
                    continue
                Numero_adivinar= int(Numero_adivinar)
                if 1 <= Numero_adivinar <= 1000:
                    break
                else:
                    print("El número debe estar entre 1 y 1000. Inténtalo de nuevo")
            except ValueError:
                print("No has introducido un número. Inténtalo de nuevo con un número del 1 al 1000")
        def Juego2():
            intentos=0
            while intentos < max_intentos:
                try:
                    Numero_insertado=input("Prueba suerte con un número del 1 al 1000: ").strip() #No es un texto
                    if not Numero_insertado:
                        print("No has introducido un número. Inténtalo de nuevo con un número del 1 al 1000.")
                        continue
                    Num_Jugador= int(Numero_insertado)
                    if Num_Jugador > Numero_adivinar:
                        intentos +=1
                        print ("Este número es mayor, pruebe con uno más pequeño")
                    elif Num_Jugador < Numero_adivinar:
                        intentos +=1
                        print ("Este número es más pequeño, pruebe con uno mayor")
                    else:
                        intentos +=1
                        print(f"Enhorabuena {Nombre2}, ¡Has ganado! El número era el {Numero_adivinar} y lo has conseguido en el intento número {intentos} de {max_intentos} posibles")
                        Registrar_jugador(Nombre2, "Ganador", intentos, nombre_dificultad)
                        Registrar_jugador(Nombre1, "Perdedor", "Te adivinaron el número", nombre_dificultad)  #Al haber adivinado el jugador 2 el número, el jugador 1 habrá perdido.
                        break
                except ValueError: #Cualquier caracter extraño
                    print("No has introducido un número. Inténtalo de nuevo con un número del 1 al 1000")
            else:
                print(f"Lo siento {Nombre2}, no has conseguido adivinar el número {Numero_adivinar}, en {max_intentos} intentos.")
                Registrar_jugador(Nombre2, "Perdedor", max_intentos, nombre_dificultad)
                Registrar_jugador(Nombre1, "Ganador", "No adivinado", nombre_dificultad)  #Al no haber adivinado, el jugador 2, el número, el jugador 1 habrá ganado.
        Juego2()
        
# Se configura las estadísticas para que recorra el excel con un "for" y nos muestre todos los valores que haya escritos:
        
    if Variable1 == 3:
        print ("Has elegido ver las estadísticas:") 
        for row in Hoja.iter_rows(min_row=1, max_row=Hoja.max_row, values_only=True):
            if any(row):
                print(row)
            print()
            
# Se configura la opción de salir del programa:
            
    if Variable1 == 4:
        print ("Has elegido Salir, te espero de vuelta")
        exit ()